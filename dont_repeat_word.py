from openpyxl import load_workbook
fn = r"words\words.xlsx"
wb = load_workbook(fn)
ws = wb["en"]
word_excel = r"words\dont_repeat_word.xlsx"
word_excel_b = load_workbook(word_excel)
word_excel_page = word_excel_b["all_word"]
count = 2
now_word = 2
let = ["A", "B", "C", "D", "E"]
num = [6403, 1666, 3309, 1564, 1319]
word_list = []
for letter in range(len(let)):
    while count != num[letter]:
        word = ws[f"{let[letter]}{count}"].value.strip()
        if word in word_list:
            print(100)
        else:
            word_excel_page[f"A{now_word}"] = word
            word_list.append(word)
            now_word += 1
        count += 1
    count = 2

wb.save(fn)
word_excel_b.save(word_excel)
wb.close()
word_excel_b.close()