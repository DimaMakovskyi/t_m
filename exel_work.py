from music import Speech
from openpyxl import load_workbook
import translator_words
from time import sleep
fn = r"words\words.xlsx"
wb = load_workbook(fn)
ws = wb["en"]
count = int(input())
let = ["A", "B", "C", "D", "E"]
num = [6403, 1666, 3309, 1564, 1319]
for letter in range(len(let)):
    while count != num[letter]:
        word = ws[f"{let[letter]}{count}"].value
        print(count)
        if "|" in word:
            ws[f"{let[letter]}{count}"] = word.replace("|", " ")
        if "@" in word:
            ws[f"{let[letter]}{count}"] = word.replace("@", " ") 
        if "?" in word:
            ws[f"{let[letter]}{count}"] = word.replace("?", " ") 
        count += 1
    count = 2


wb.save(fn)
wb.close()