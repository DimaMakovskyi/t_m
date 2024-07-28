from music import Speech
from time import sleep
from openpyxl import load_workbook
fn = r"words\words.xlsx"
wb = load_workbook(fn)
ws = wb["en"]
count = int(input())
create_music = Speech()
for letter in ["E"]:
    while count != 1319:
        word = ws[f"{letter}{count}"].value
        create_music.create_speech(word)
        print(count)
        count += 1
        sleep(1)


wb.close()